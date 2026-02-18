"""
Reporting workers for PDF and Excel generation.
"""
import asyncio
import json
import io
from datetime import datetime
from typing import Dict, Any, Optional

from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    PageBreak,
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from app.workers.celery_app import celery_app
from app.core.database import AsyncSessionLocal
from app.core.logging import get_logger
from app.core.minio_client import upload_report
from app.models.report import Report, ReportStatus
from app.models.analytics_job import AnalyticsJob
from app.services.report_data import get_report_data
from sqlalchemy import select, update


logger = get_logger(__name__)


def generate_pdf(report: Report, data: Dict[str, Any], analytics_results: Optional[Dict] = None) -> bytes:
    """
    Generate PDF report.
    
    Args:
        report: Report model instance
        data: Report data from get_report_data()
        analytics_results: Optional analytics results
    
    Returns:
        PDF file as bytes
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1e40af'),
        alignment=TA_CENTER,
        spaceAfter=30,
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=16,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=12,
    )
    
    story = []
    
    # Page 1: Cover
    story.append(Spacer(1, 3*cm))
    story.append(Paragraph(report.title or "Factory Operations Report", title_style))
    story.append(Spacer(1, 1*cm))
    
    cover_data = [
        ["Date Range:", f"{report.date_range_start.strftime('%Y-%m-%d')} to {report.date_range_end.strftime('%Y-%m-%d')}"],
        ["Generated:", datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')],
        ["Devices:", str(len(data['devices']))],
        ["Alerts:", str(len(data['alerts']))],
    ]
    
    cover_table = Table(cover_data, colWidths=[5*cm, 10*cm])
    cover_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#4b5563')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
    ]))
    story.append(cover_table)
    story.append(PageBreak())
    
    # Page 2: Executive Summary
    story.append(Paragraph("Executive Summary", heading_style))
    story.append(Spacer(1, 0.5*cm))
    
    alert_summary = data.get('alert_summary', {})
    summary_data = [
        ["Metric", "Value"],
        ["Total Devices", str(len(data['devices']))],
        ["Total Alerts", str(len(data['alerts']))],
        ["Critical Alerts", str(alert_summary.get('critical', 0))],
        ["High Alerts", str(alert_summary.get('high', 0))],
        ["Medium Alerts", str(alert_summary.get('medium', 0))],
        ["Low Alerts", str(alert_summary.get('low', 0))],
    ]
    
    summary_table = Table(summary_data, colWidths=[8*cm, 6*cm])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')]),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
    ]))
    story.append(summary_table)
    story.append(PageBreak())
    
    # Page 3: Device Details
    story.append(Paragraph("Device Details", heading_style))
    story.append(Spacer(1, 0.5*cm))
    
    for device in data['devices']:
        story.append(Paragraph(f"<b>{device['name'] or device['device_key']}</b>", styles['Heading3']))
        
        device_info = [
            ["Device Key:", device['device_key']],
            ["Region:", device.get('region', 'N/A')],
            ["Manufacturer:", device.get('manufacturer', 'N/A')],
            ["Model:", device.get('model', 'N/A')],
        ]
        
        device_table = Table(device_info, colWidths=[4*cm, 10*cm])
        device_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(device_table)
        
        # Telemetry statistics for this device
        telemetry = data.get('telemetry_summary', {}).get(str(device['id']), {})
        if telemetry:
            story.append(Spacer(1, 0.3*cm))
            story.append(Paragraph("<b>Telemetry Statistics:</b>", styles['Normal']))
            
            telemetry_data = [["Parameter", "Min", "Max", "Average"]]
            for param, stats in telemetry.items():
                telemetry_data.append([
                    param,
                    f"{stats['min']:.2f}",
                    f"{stats['max']:.2f}",
                    f"{stats['avg']:.2f}",
                ])
            
            telemetry_table = Table(telemetry_data, colWidths=[4*cm, 3*cm, 3*cm, 3*cm])
            telemetry_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6366f1')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')]),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
            ]))
            story.append(telemetry_table)
        
        story.append(Spacer(1, 0.5*cm))
    
    if data['devices']:
        story.append(PageBreak())
    
    # Page 4: Alerts Log
    if data['alerts']:
        story.append(Paragraph("Alerts Log", heading_style))
        story.append(Spacer(1, 0.5*cm))
        
        alert_data = [["Timestamp", "Severity", "Device ID", "Message"]]
        for alert in data['alerts'][:50]:  # Limit to 50 alerts
            alert_data.append([
                alert['triggered_at'][:19],  # Remove timezone
                alert['severity'].upper(),
                str(alert['device_id']),
                alert['message'][:60],  # Truncate long messages
            ])
        
        alert_table = Table(alert_data, colWidths=[4*cm, 2.5*cm, 2.5*cm, 6*cm])
        alert_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#dc2626')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#fef2f2')]),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(alert_table)
        story.append(PageBreak())
    
    # Page 5+: Analytics Results
    if analytics_results:
        story.append(Paragraph("Analytics Results", heading_style))
        story.append(Spacer(1, 0.5*cm))
        
        summary_text = analytics_results.get('summary', 'No summary available')
        story.append(Paragraph(f"<b>Summary:</b> {summary_text}", styles['Normal']))
        story.append(Spacer(1, 0.3*cm))
        
        # Show mode and models used
        if 'mode' in analytics_results:
            story.append(Paragraph(f"<b>Mode:</b> {analytics_results['mode']}", styles['Normal']))
        if 'models_used' in analytics_results:
            story.append(Paragraph(f"<b>Models:</b> {', '.join(analytics_results['models_used'])}", styles['Normal']))
        
        story.append(Spacer(1, 0.5*cm))
        
        # Detailed results
        results = analytics_results.get('results', {})
        
        if 'anomaly' in results:
            anomaly = results['anomaly']
            story.append(Paragraph("<b>Anomaly Detection:</b>", styles['Heading3']))
            story.append(Paragraph(f"Anomaly Count: {anomaly.get('anomaly_count', 0)}", styles['Normal']))
            story.append(Paragraph(f"Anomaly Score: {anomaly.get('anomaly_score', 0):.2%}", styles['Normal']))
            story.append(Spacer(1, 0.3*cm))
        
        if 'forecast' in results:
            forecast = results['forecast']
            story.append(Paragraph("<b>Energy Forecast:</b>", styles['Heading3']))
            story.append(Paragraph(f"Horizon: {forecast.get('horizon_days', 0)} days", styles['Normal']))
            story.append(Paragraph(f"Forecast Points: {forecast.get('forecast_points', 0)}", styles['Normal']))
            story.append(Spacer(1, 0.3*cm))
        
        if 'failure' in results:
            failure = results['failure']
            story.append(Paragraph("<b>Failure Prediction:</b>", styles['Heading3']))
            story.append(Paragraph(f"Failure Probability: {failure.get('failure_probability', 0):.2%}", styles['Normal']))
            story.append(Paragraph(f"Risk Level: {failure.get('risk_level', 'unknown').upper()}", styles['Normal']))
    
    # Build PDF
    doc.build(story)
    
    pdf_bytes = buffer.getvalue()
    buffer.close()
    
    return pdf_bytes


def generate_excel(report: Report, data: Dict[str, Any], analytics_results: Optional[Dict] = None) -> bytes:
    """
    Generate Excel report.
    
    Args:
        report: Report model instance
        data: Report data from get_report_data()
        analytics_results: Optional analytics results
    
    Returns:
        Excel file as bytes
    """
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Sheet 1: Summary
    ws_summary = wb.create_sheet("Summary")
    ws_summary.append(["Factory Operations Report"])
    ws_summary.append([])
    ws_summary.append(["Report Title", report.title or "Factory Operations Report"])
    ws_summary.append(["Date Range", f"{report.date_range_start.strftime('%Y-%m-%d')} to {report.date_range_end.strftime('%Y-%m-%d')}"])
    ws_summary.append(["Generated", datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')])
    ws_summary.append([])
    ws_summary.append(["Metric", "Value"])
    ws_summary.append(["Total Devices", len(data['devices'])])
    ws_summary.append(["Total Alerts", len(data['alerts'])])
    
    alert_summary = data.get('alert_summary', {})
    ws_summary.append(["Critical Alerts", alert_summary.get('critical', 0)])
    ws_summary.append(["High Alerts", alert_summary.get('high', 0)])
    ws_summary.append(["Medium Alerts", alert_summary.get('medium', 0)])
    ws_summary.append(["Low Alerts", alert_summary.get('low', 0)])
    
    # Style summary sheet
    ws_summary['A1'].font = Font(size=16, bold=True, color="1e40af")
    for row in ws_summary['A7:B13']:
        for cell in row:
            cell.font = Font(bold=True if cell.column == 1 else False)
            cell.fill = PatternFill(start_color="e5e7eb", end_color="e5e7eb", fill_type="solid") if cell.row == 7 else PatternFill()
    
    # Sheet 2: Devices
    ws_devices = wb.create_sheet("Devices")
    ws_devices.append(["Device ID", "Name", "Device Key", "Region", "Manufacturer", "Model", "Last Seen"])
    
    for device in data['devices']:
        ws_devices.append([
            device['id'],
            device['name'],
            device['device_key'],
            device.get('region', ''),
            device.get('manufacturer', ''),
            device.get('model', ''),
            device.get('last_seen', ''),
        ])
    
    # Style devices header
    for cell in ws_devices[1]:
        cell.font = Font(bold=True, color="ffffff")
        cell.fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
        cell.alignment = Alignment(horizontal="left")
    
    # Sheet 3: Alerts
    ws_alerts = wb.create_sheet("Alerts")
    ws_alerts.append(["Alert ID", "Device ID", "Severity", "Message", "Triggered At", "Resolved At"])
    
    for alert in data['alerts']:
        ws_alerts.append([
            alert['id'],
            alert['device_id'],
            alert['severity'].upper(),
            alert['message'],
            alert['triggered_at'],
            alert.get('resolved_at', ''),
        ])
    
    # Style alerts header
    for cell in ws_alerts[1]:
        cell.font = Font(bold=True, color="ffffff")
        cell.fill = PatternFill(start_color="dc2626", end_color="dc2626", fill_type="solid")
        cell.alignment = Alignment(horizontal="left")
    
    # Sheet 4: Telemetry Summary
    ws_telemetry = wb.create_sheet("Telemetry")
    ws_telemetry.append(["Device ID", "Parameter", "Min", "Max", "Average"])
    
    for device_id, parameters in data.get('telemetry_summary', {}).items():
        for param, stats in parameters.items():
            ws_telemetry.append([
                device_id,
                param,
                round(stats['min'], 2),
                round(stats['max'], 2),
                round(stats['avg'], 2),
            ])
    
    # Style telemetry header
    for cell in ws_telemetry[1]:
        cell.font = Font(bold=True, color="ffffff")
        cell.fill = PatternFill(start_color="6366f1", end_color="6366f1", fill_type="solid")
        cell.alignment = Alignment(horizontal="left")
    
    # Sheet 5: Analytics (if included)
    if analytics_results:
        ws_analytics = wb.create_sheet("Analytics")
        ws_analytics.append(["Analytics Results"])
        ws_analytics.append([])
        ws_analytics.append(["Summary", analytics_results.get('summary', 'No summary')])
        ws_analytics.append(["Mode", analytics_results.get('mode', 'N/A')])
        ws_analytics.append(["Models Used", ', '.join(analytics_results.get('models_used', []))])
        ws_analytics.append([])
        
        results = analytics_results.get('results', {})
        
        if 'anomaly' in results:
            anomaly = results['anomaly']
            ws_analytics.append(["Anomaly Detection"])
            ws_analytics.append(["Anomaly Count", anomaly.get('anomaly_count', 0)])
            ws_analytics.append(["Anomaly Score", f"{anomaly.get('anomaly_score', 0):.2%}"])
            ws_analytics.append([])
        
        if 'forecast' in results:
            forecast = results['forecast']
            ws_analytics.append(["Energy Forecast"])
            ws_analytics.append(["Horizon Days", forecast.get('horizon_days', 0)])
            ws_analytics.append(["Forecast Points", forecast.get('forecast_points', 0)])
            ws_analytics.append([])
        
        if 'failure' in results:
            failure = results['failure']
            ws_analytics.append(["Failure Prediction"])
            ws_analytics.append(["Failure Probability", f"{failure.get('failure_probability', 0):.2%}"])
            ws_analytics.append(["Risk Level", failure.get('risk_level', 'unknown').upper()])
        
        ws_analytics['A1'].font = Font(size=14, bold=True, color="1e40af")
    
    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    excel_bytes = buffer.getvalue()
    buffer.close()
    
    return excel_bytes


# Synchronous helper functions for Celery
def get_report_sync(report_id: str) -> Report:
    """Get report from database synchronously."""
    async def _get():
        async with AsyncSessionLocal() as db:
            result = await db.execute(
                select(Report).where(Report.id == report_id)
            )
            return result.scalar_one()
    
    return asyncio.run(_get())


def get_analytics_results_sync(job_id: str) -> Optional[Dict[str, Any]]:
    """Get analytics results synchronously."""
    # This would fetch from MinIO or database
    # For now, return None if not available
    # TODO: Implement fetching analytics results from MinIO
    return None


def update_report_status_sync(
    report_id: str,
    status: str,
    file_url: str = None,
    file_size: int = None,
    error: str = None,
) -> None:
    """Update report status synchronously."""
    async def _update():
        async with AsyncSessionLocal() as db:
            update_data = {"status": ReportStatus[status.upper()]}
            
            if file_url:
                update_data["file_url"] = file_url
            
            if file_size:
                update_data["file_size_bytes"] = file_size
            
            if error:
                update_data["error_message"] = error
            
            # Set expiration (24 hours from now)
            if status == "complete":
                from datetime import timedelta
                update_data["expires_at"] = datetime.utcnow() + timedelta(hours=24)
            
            await db.execute(
                update(Report)
                .where(Report.id == report_id)
                .values(**update_data)
            )
            await db.commit()
    
    asyncio.run(_update())


@celery_app.task(name="generate_report", bind=True, max_retries=1, queue="reporting")
def generate_report_task(self, report_id: str):
    """
    Generate report asynchronously.
    
    Args:
        report_id: Report ID
    """
    logger.info("report.start", report_id=report_id)
    
    # Update status to running
    update_report_status_sync(report_id, "running")
    
    try:
        # Fetch report details
        report = get_report_sync(report_id)
        
        logger.info(
            "report.fetching_data",
            report_id=report_id,
            factory_id=report.factory_id,
            format=report.format.value,
            device_count=len(report.device_ids),
        )
        
        # Fetch report data
        data = asyncio.run(
            get_report_data(
                report.factory_id,
                report.device_ids,
                report.date_range_start,
                report.date_range_end,
            )
        )
        
        logger.info(
            "report.data_fetched",
            report_id=report_id,
            devices=len(data['devices']),
            alerts=len(data['alerts']),
        )
        
        # Fetch analytics results if requested
        analytics = None
        if report.include_analytics and report.analytics_job_id:
            analytics = get_analytics_results_sync(report.analytics_job_id)
        
        # Generate file based on format
        if report.format.value == "pdf":
            file_bytes = generate_pdf(report, data, analytics)
        elif report.format.value == "excel":
            file_bytes = generate_excel(report, data, analytics)
        else:  # json
            file_bytes = json.dumps({**data, "analytics": analytics}, indent=2, default=str).encode()
        
        logger.info(
            "report.file_generated",
            report_id=report_id,
            format=report.format.value,
            size_bytes=len(file_bytes),
        )
        
        # Upload to MinIO
        file_url = upload_report(report.factory_id, report_id, file_bytes, report.format.value)
        
        # Update report status
        update_report_status_sync(report_id, "complete", file_url=file_url, file_size=len(file_bytes))
        
        logger.info(
            "report.success",
            report_id=report_id,
            file_url=file_url,
        )
        
        return {"status": "complete", "file_url": file_url}
        
    except Exception as e:
        logger.error(
            "report.failed",
            report_id=report_id,
            error=str(e),
            exc_info=True,
        )
        
        # Update report status to failed
        update_report_status_sync(report_id, "failed", error=str(e))
        
        # Retry if possible
        raise self.retry(exc=e)
